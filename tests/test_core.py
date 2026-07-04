import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from datafoldit import db
from datafoldit.excel_io import DEFAULT_SOURCE_XLSX, export_report_workbook, import_company_workbook
from datafoldit.invoice_pdf import parse_invoice_text
from datafoldit.paystub_import import extract_paystub_rows_from_file, parse_paystub_text
from datafoldit.transaction_import import parse_transaction_text
from datafoldit.web import (
    add_bank_transaction_batch,
    add_invoice_batch,
    add_payroll_batch,
    attachment_link,
    filter_bank_rows,
    filter_expense_rows,
    filter_invoice_rows,
    filter_payroll_rows,
    filter_rows_by_period,
    render_bank,
    render_bank_edit,
    render_expense_edit,
    render_expenses,
    render_invoice_bulk_review,
    render_invoice_edit,
    render_invoices,
    render_payroll,
    render_paystub_bulk_review,
    render_paystub_review,
    render_transaction_bulk_review,
)


class DataFoldCoreTests(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.tmpdir.name) / "test.sqlite"
        self.conn = db.connect(self.db_path)
        db.init_db(self.conn)

    def tearDown(self):
        self.conn.close()
        self.tmpdir.cleanup()

    def test_import_current_workbook(self):
        if not DEFAULT_SOURCE_XLSX.exists():
            self.skipTest(f"Missing source workbook: {DEFAULT_SOURCE_XLSX}")
        counts = import_company_workbook(self.conn, DEFAULT_SOURCE_XLSX, replace=True)
        self.assertEqual(counts["expenses"], 25)
        self.assertEqual(counts["bank_transactions"], 22)
        self.assertEqual(counts["payroll_entries"], 6)
        self.assertEqual(counts["invoices"], 13)
        self.assertAlmostEqual(db.current_balance(self.conn), 10334.58, places=2)
        monthly = db.monthly_bank_summary(self.conn)
        self.assertEqual(monthly[0]["month"], "2025-12")
        self.assertAlmostEqual(monthly[0]["opening"], 6400.00, places=2)
        self.assertAlmostEqual(monthly[0]["closing"], 6200.00, places=2)

    def test_report_export_is_valid_xlsx(self):
        if not DEFAULT_SOURCE_XLSX.exists():
            self.skipTest(f"Missing source workbook: {DEFAULT_SOURCE_XLSX}")
        import_company_workbook(self.conn, DEFAULT_SOURCE_XLSX, replace=True)
        output = Path(self.tmpdir.name) / "report.xlsx"
        export_report_workbook(self.conn, output, period="all")
        workbook = load_workbook(output, data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["Summary", "Bank Transactions", "Expenses", "Payroll", "Invoices"],
        )
        self.assertEqual(workbook["Summary"]["A1"].value, "DataFold IT Operations Report")
        self.assertAlmostEqual(float(workbook["Summary"]["B5"].value), 10334.58, places=2)

    def test_period_filter_supports_year_and_blank_month(self):
        if not DEFAULT_SOURCE_XLSX.exists():
            self.skipTest(f"Missing source workbook: {DEFAULT_SOURCE_XLSX}")
        import_company_workbook(self.conn, DEFAULT_SOURCE_XLSX, replace=True)
        rows = db.rows_for_table(self.conn, "bank_transactions")
        may_rows = filter_rows_by_period(rows, "date", {"year": "2026", "month": "05"})
        year_rows = filter_rows_by_period(rows, "date", {"year": "2026", "month": ""})
        all_rows = filter_rows_by_period(rows, "date", {"year": "", "month": ""})
        self.assertEqual(len(may_rows), 6)
        self.assertEqual(len(year_rows), 21)
        self.assertEqual(len(all_rows), 22)

    def test_smart_transaction_text_parser(self):
        text = """
        ACH CREDIT
        Payment received
        From: Acme Client LLC
        Transaction Date: May 28, 2026
        Amount: $2,500.00
        """
        parsed = parse_transaction_text(text, "payment.pdf")
        self.assertEqual(parsed["date"], "2026-05-28")
        self.assertEqual(parsed["type"], "Deposit")
        self.assertEqual(parsed["detail"], "Acme Client LLC")
        self.assertEqual(parsed["category"], "Client Payment")
        self.assertEqual(parsed["source"], "ACH")
        self.assertAlmostEqual(parsed["amount"], 2500.00, places=2)

    def test_expense_and_payroll_attachment_paths_are_saved(self):
        db.add_expense(
            self.conn,
            {
                "date": "2026-05-28",
                "amount": "42.50",
                "vendor": "Zoom",
                "attachment_path": "/tmp/zoom-receipt.pdf",
            },
        )
        db.add_payroll_entry(
            self.conn,
            {
                "month": "2026-05",
                "first_name": "Vamsi",
                "gross": "1000",
                "attachment_path": "/tmp/payroll.pdf",
            },
        )
        expense = db.rows_for_table(self.conn, "expenses")[0]
        payroll = db.rows_for_table(self.conn, "payroll_entries")[0]
        self.assertEqual(expense["attachment_path"], "/tmp/zoom-receipt.pdf")
        self.assertEqual(payroll["attachment_path"], "/tmp/payroll.pdf")

    def test_multiple_attachment_paths_render_as_multiple_links(self):
        html = attachment_link("/tmp/receipt-one.pdf\n/tmp/receipt-two.docx")
        self.assertIn("receipt-one.pdf", html)
        self.assertIn("receipt-two.docx", html)

    def test_bank_and_expense_filters_and_delete_helpers(self):
        bank_id = db.add_bank_transaction(
            self.conn,
            {
                "date": "2026-05-28",
                "type": "Expense",
                "detail": "Laptop",
                "source": "Vamsi",
                "amount": "900",
            },
        )
        db.add_bank_transaction(
            self.conn,
            {
                "date": "2026-05-28",
                "type": "Expense",
                "detail": "Software",
                "source": "Aditya",
                "amount": "100",
            },
        )
        db.add_expense(
            self.conn,
            {
                "date": "2026-05-28",
                "vendor": "Apple",
                "paid_by": "Vamsi",
                "amount": "900",
            },
        )
        expense_id = db.add_expense(
            self.conn,
            {
                "date": "2026-05-28",
                "vendor": "Adobe",
                "paid_by": "Aditya",
                "amount": "100",
            },
        )
        bank_rows = filter_bank_rows(db.rows_for_table(self.conn, "bank_transactions"), {"source": "Vamsi", "year": "", "month": ""})
        expense_rows = filter_expense_rows(db.rows_for_table(self.conn, "expenses"), {"paid_by": "Aditya", "year": "", "month": ""})
        self.assertEqual([row["detail"] for row in bank_rows], ["Laptop"])
        self.assertEqual([row["vendor"] for row in expense_rows], ["Adobe"])
        db.delete_bank_transaction(self.conn, bank_id)
        db.delete_expense(self.conn, expense_id)
        remaining_bank = self.conn.execute("SELECT COUNT(*) AS count FROM bank_transactions WHERE id = ?", (bank_id,)).fetchone()
        remaining_expense = self.conn.execute("SELECT COUNT(*) AS count FROM expenses WHERE id = ?", (expense_id,)).fetchone()
        self.assertEqual(remaining_bank["count"], 0)
        self.assertEqual(remaining_expense["count"], 0)

    def test_bank_expense_and_invoice_edit_forms_and_updates(self):
        bank_id = db.add_bank_transaction(
            self.conn,
            {
                "date": "2026-05-28",
                "type": "Expense",
                "detail": "Original laptop",
                "source": "Vamsi",
                "amount": "900",
                "attachment_path": "/tmp/original-bank.pdf",
            },
        )
        expense_id = db.add_expense(
            self.conn,
            {
                "date": "2026-05-28",
                "vendor": "Original Vendor",
                "paid_by": "Aditya",
                "amount": "100",
            },
        )
        invoice_id = db.add_invoice(
            self.conn,
            {
                "date": "2026-05-28",
                "invoice_number": "INV-EDIT-001",
                "customer": "Original Customer",
                "amount": "1000",
                "due_date": "2026-06-01",
                "status": "Not Received",
            },
        )
        payroll_id = db.add_payroll_entry(
            self.conn,
            {
                "month": "2026-05",
                "first_name": "Original",
                "last_name": "Employee",
                "vendor": "Original Vendor",
                "client": "Original Client",
                "job_start": "2026-05-01",
                "job_end": "2026-05-31",
                "vendor_pay": "60",
                "pct": "30",
                "hours": "10",
                "attachment_path": "/tmp/original-paystub.pdf",
                "paystub_sent": "N",
            },
        )
        self.assertIn('action="/bank/update"', render_bank_edit(self.conn, bank_id))
        self.assertIn('action="/expenses/update"', render_expense_edit(self.conn, expense_id))
        self.assertIn('action="/invoices/update"', render_invoice_edit(self.conn, invoice_id))
        db.update_bank_transaction(
            self.conn,
            bank_id,
            {"date": "2026-06-01", "type": "Deposit", "detail": "Updated payment", "source": "ACH", "amount": "1250"},
        )
        db.update_expense(
            self.conn,
            expense_id,
            {"date": "2026-06-02", "vendor": "Updated Vendor", "paid_by": "Vamsi", "amount": "250"},
        )
        db.update_invoice(
            self.conn,
            invoice_id,
            {
                "date": "2026-06-03",
                "invoice_number": "INV-EDIT-002",
                "customer": "Updated Customer",
                "amount": "1200",
                "due_date": "2026-06-30",
                "balance_due": "1200",
                "status": "Not Received",
            },
        )
        db.update_payroll_entry(
            self.conn,
            payroll_id,
            {
                "month": "2026-06",
                "first_name": "Updated",
                "last_name": "Employee",
                "vendor": "Updated Vendor",
                "client": "Updated Client",
                "job_start": "2026-05-01",
                "job_end": "2026-05-31",
                "vendor_pay": "60",
                "hours": "20",
                "gross": "1200",
                "tax": "120",
                "credit_date": "2026-06-10",
                "paystub_sent": "Yes",
            },
        )
        bank = self.conn.execute("SELECT date, type, detail, amount, attachment_path FROM bank_transactions WHERE id = ?", (bank_id,)).fetchone()
        expense = self.conn.execute("SELECT date, vendor, paid_by, amount FROM expenses WHERE id = ?", (expense_id,)).fetchone()
        invoice = self.conn.execute("SELECT date, invoice_number, customer, amount, balance_due FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
        payroll = self.conn.execute("SELECT month, first_name, last_name, vendor, client, hours, gross, tax, commission, employee_pay, credit_date, attachment_path, paystub_sent FROM payroll_entries WHERE id = ?", (payroll_id,)).fetchone()
        self.assertEqual(bank["date"], "2026-06-01")
        self.assertEqual(bank["type"], "Deposit")
        self.assertEqual(bank["detail"], "Updated payment")
        self.assertAlmostEqual(bank["amount"], 1250.0)
        self.assertEqual(bank["attachment_path"], "/tmp/original-bank.pdf")
        self.assertEqual(expense["vendor"], "Updated Vendor")
        self.assertEqual(expense["paid_by"], "Vamsi")
        self.assertAlmostEqual(expense["amount"], 250.0)
        self.assertEqual(invoice["invoice_number"], "INV-EDIT-002")
        self.assertEqual(invoice["customer"], "Updated Customer")
        self.assertAlmostEqual(invoice["amount"], 1200.0)
        self.assertAlmostEqual(invoice["balance_due"], 1200.0)
        self.assertEqual(payroll["month"], "2026-06")
        self.assertEqual(payroll["first_name"], "Updated")
        self.assertEqual(payroll["client"], "Updated Client")
        self.assertAlmostEqual(payroll["hours"], 20.0)
        self.assertAlmostEqual(payroll["tax"], 120.0)
        self.assertAlmostEqual(payroll["commission"], 0.0)
        self.assertAlmostEqual(payroll["employee_pay"], 1080.0)
        self.assertEqual(payroll["attachment_path"], "/tmp/original-paystub.pdf")
        self.assertEqual(payroll["paystub_sent"], "Y")

    def test_payroll_filters_by_candidate_name(self):
        db.add_payroll_entry(
            self.conn,
            {
                "month": "2026-05",
                "first_name": "Alex",
                "last_name": "Rao",
                "client": "Acme Client",
                "hours": "10",
                "gross": "600",
            },
        )
        db.add_payroll_entry(
            self.conn,
            {
                "month": "2026-05",
                "first_name": "Maya",
                "last_name": "Patel",
                "client": "Beta Client",
                "hours": "8",
                "gross": "480",
            },
        )
        rows = db.rows_for_table(self.conn, "payroll_entries")
        alex_rows = filter_payroll_rows(rows, {"candidate": "Alex Rao", "year": "", "month": ""})
        beta_rows = filter_payroll_rows(rows, {"candidate": "Maya Patel", "year": "", "month": ""})
        self.assertEqual([row["client"] for row in alex_rows], ["Acme Client"])
        self.assertEqual([row["first_name"] for row in beta_rows], ["Maya"])

    def test_payroll_calculates_from_rate_hours_and_tax(self):
        db.add_payroll_entry(
            self.conn,
            {
                "month": "2026-05",
                "first_name": "Alex",
                "last_name": "Rao",
                "vendor_pay": "60",
                "hours": "10",
                "tax": "75",
            },
        )
        row = db.rows_for_table(self.conn, "payroll_entries")[0]
        self.assertAlmostEqual(row["gross"], 600.00, places=2)
        self.assertAlmostEqual(row["tax"], 75.00, places=2)
        self.assertAlmostEqual(row["commission"], 0.00, places=2)
        self.assertAlmostEqual(row["employee_pay"], 525.00, places=2)

    def test_invoice_ocr_text_parser(self):
        text = """
        Invoice
        Zoom Communications, Inc.
        Invoice Date:
        Invoice #:
        Due Date:
        Jan 13, 2025
        INV288750792
        Jan 13, 2025
        Bill To Address:
        A2systemsLLC
        Subtotal $15.00
        Total (Including Taxes, Fees & Surcharges) $17.21
        """
        parsed = parse_invoice_text(text)
        self.assertEqual(parsed["invoice_number"], "INV288750792")
        self.assertEqual(parsed["date"], "2025-01-13")
        self.assertEqual(parsed["due_date"], "2025-01-13")
        self.assertEqual(parsed["customer"], "A2systemsLLC")
        self.assertAlmostEqual(parsed["amount"], 17.21, places=2)
        inline_customer = parse_invoice_text(
            """
            Invoice Number: INV-TEST-777
            Invoice Date: June 25, 2026
            Customer: Example Client LLC
            Due Date: July 25, 2026
            Total Amount: $987.65
            Balance Due: $987.65
            """
        )
        self.assertEqual(inline_customer["customer"], "Example Client LLC")

    def test_invoice_parser_handles_day_month_dates_and_paid_balance(self):
        text = """
        Invoice
        # INV-000001
        Bill To
        The Quantum Core Technologies LLC
        Invoice Date: 12 Feb 2026
        Terms : Net 60
        Due Date : 13 Apr 2026
        Total $6,160.00
        Payment Made (-) 6,160.00
        Balance Due $0.00
        """
        parsed = parse_invoice_text(text)
        self.assertEqual(parsed["invoice_number"], "INV-000001")
        self.assertEqual(parsed["date"], "2026-02-12")
        self.assertEqual(parsed["due_date"], "2026-04-13")
        self.assertEqual(parsed["customer"], "The Quantum Core Technologies LLC")
        self.assertEqual(parsed["status"], "Paid")
        self.assertEqual(parsed["received"], "Y")
        self.assertAlmostEqual(parsed["amount"], 6160.00, places=2)
        self.assertAlmostEqual(parsed["balance_due"], 0.00, places=2)

    def test_invoice_parser_handles_screenshot_ocr_layout(self):
        text = """
        Invoice Details
        Balance Due
        $11,440.00
        Bansar Technologies Inc .
        Due on: Thu, Jun 25 2026
        Terms: Net 35
        Invoice# Invoice Date
        INV-000012 Thu, May 21 2026
        Consulting Service $11,440.00
        Subtotal $11,440.00
        Total $11,440.00
        """
        parsed = parse_invoice_text(text)
        self.assertEqual(parsed["invoice_number"], "INV-000012")
        self.assertEqual(parsed["date"], "2026-05-21")
        self.assertEqual(parsed["due_date"], "2026-06-25")
        self.assertEqual(parsed["customer"], "Bansar Technologies Inc")
        self.assertEqual(parsed["status"], "Open")
        self.assertAlmostEqual(parsed["amount"], 11440.00, places=2)
        self.assertAlmostEqual(parsed["balance_due"], 11440.00, places=2)

    def test_invoice_status_dropdown_values_map_to_received_and_void(self):
        db.add_invoice(
            self.conn,
            {
                "date": "2026-05-29",
                "invoice_number": "INV-STATUS-001",
                "customer": "Acme LLC",
                "amount": "100",
                "status": "Received",
            },
        )
        db.add_invoice(
            self.conn,
            {
                "date": "2026-05-29",
                "invoice_number": "INV-STATUS-002",
                "customer": "Beta LLC",
                "amount": "200",
                "status": "Not Received",
            },
        )
        db.add_invoice(
            self.conn,
            {
                "date": "2026-05-29",
                "invoice_number": "INV-STATUS-003",
                "customer": "Void LLC",
                "amount": "300",
                "status": "Void",
            },
        )
        rows = {
            row["invoice_number"]: row
            for row in self.conn.execute("SELECT * FROM invoices WHERE invoice_number LIKE 'INV-STATUS-%'")
        }
        acme_rows = filter_invoice_rows(list(rows.values()), {"customer": "Acme LLC", "status": "", "year": "", "month": ""})
        received_rows = filter_invoice_rows(list(rows.values()), {"customer": "", "status": "Received", "year": "", "month": ""})
        self.assertEqual([row["invoice_number"] for row in acme_rows], ["INV-STATUS-001"])
        self.assertEqual([row["invoice_number"] for row in received_rows], ["INV-STATUS-001"])
        self.assertEqual(rows["INV-STATUS-001"]["received"], "Y")
        self.assertEqual(rows["INV-STATUS-001"]["is_void"], 0)
        self.assertEqual(rows["INV-STATUS-001"]["status"], "Paid")
        self.assertEqual(rows["INV-STATUS-002"]["received"], "N")
        self.assertEqual(rows["INV-STATUS-002"]["is_void"], 0)
        self.assertEqual(rows["INV-STATUS-002"]["status"], "Open")
        self.assertEqual(rows["INV-STATUS-003"]["received"], "N")
        self.assertEqual(rows["INV-STATUS-003"]["is_void"], 1)
        self.assertEqual(rows["INV-STATUS-003"]["status"], "VOID")
        db.update_invoice_status(self.conn, rows["INV-STATUS-002"]["id"], "Received")
        updated = self.conn.execute(
            "SELECT status, received, is_void, balance_due FROM invoices WHERE invoice_number = ?",
            ("INV-STATUS-002",),
        ).fetchone()
        self.assertEqual(updated["status"], "Paid")
        self.assertEqual(updated["received"], "Y")
        self.assertEqual(updated["is_void"], 0)
        self.assertAlmostEqual(updated["balance_due"], 0.0, places=2)
        db.update_invoice_status(self.conn, rows["INV-STATUS-002"]["id"], "Void")
        voided = self.conn.execute(
            "SELECT status, received, is_void, balance_due FROM invoices WHERE invoice_number = ?",
            ("INV-STATUS-002",),
        ).fetchone()
        self.assertEqual(voided["status"], "VOID")
        self.assertEqual(voided["received"], "N")
        self.assertEqual(voided["is_void"], 1)
        self.assertAlmostEqual(voided["balance_due"], 0.0, places=2)
        db.delete_invoice(self.conn, rows["INV-STATUS-002"]["id"])
        deleted = self.conn.execute(
            "SELECT COUNT(*) AS count FROM invoices WHERE invoice_number = ?",
            ("INV-STATUS-002",),
        ).fetchone()
        self.assertEqual(deleted["count"], 0)

    def test_invoice_upload_accepts_multiple_file_types(self):
        db.add_invoice(
            self.conn,
            {
                "date": "2026-05-29",
                "invoice_number": "INV-CLIENT-001",
                "customer": "Dropdown Client LLC",
                "amount": "100",
            },
        )
        html = render_invoices(self.conn)
        self.assertIn('name="attachment" form="invoice-create-form"', html)
        self.assertIn('data-extract-kind="invoice"', html)
        self.assertIn('data-extract-url="/invoices/extract-inline"', html)
        self.assertIn('id="invoice-create-row"', html)
        self.assertIn('data-inline-create-toggle', html)
        self.assertNotIn("Read Invoice Files", html)
        self.assertNotIn("New Invoice", html)
        self.assertNotIn("accept=", html)
        self.assertIn('class="grid cols-4"', html)
        self.assertIn("Total Invoice", html)
        self.assertIn("Received", html)
        self.assertIn("Commission Received", html)
        self.assertIn("Outstanding", html)
        self.assertIn("Overdue", html)
        self.assertNotIn("Next #", html)
        self.assertIn('<datalist id="customer-options">', html)
        self.assertIn("Dropdown Client LLC", html)

    def test_invoice_ledger_has_sort_headers_overdue_status_and_edit_action(self):
        db.add_invoice(
            self.conn,
            {
                "date": "2026-05-01",
                "invoice_number": "INV-SORT-001",
                "customer": "Overdue Client",
                "amount": "100",
                "due_date": "2020-01-01",
                "balance_due": "100",
                "status": "Not Received",
            },
        )
        db.add_invoice(
            self.conn,
            {
                "date": "2026-05-02",
                "invoice_number": "INV-SORT-002",
                "customer": "Paid Client",
                "amount": "200",
                "due_date": "2020-01-01",
                "status": "Received",
            },
        )
        html = render_invoices(self.conn, filters={"year": "", "month": "", "customer": "", "status": "", "sort": "amount", "direction": "asc"})
        self.assertIn("Due Status", html)
        self.assertIn("Overdue", html)
        self.assertIn("$100.00", html)
        self.assertIn("Commission %", html)
        self.assertIn("Commission Amount", html)
        self.assertIn('sort=commission_amount', html)
        self.assertIn('sort=balance_due', html)
        self.assertIn('sort=invoice_number', html)
        self.assertIn('id="invoice-row-form-', html)
        self.assertIn('action="/invoices/update"', html)
        self.assertIn('class="cell-editor"', html)
        self.assertIn('inline-save-button', html)
        self.assertNotIn('/invoices/edit?id=', html)
        self.assertIn('edit-icon-button', html)

    def test_bank_expense_payroll_pages_have_bulk_upload_and_delete_controls(self):
        db.add_bank_transaction(
            self.conn,
            {
                "date": "2026-05-29",
                "type": "Expense",
                "detail": "Cloud",
                "source": "Vamsi",
                "amount": "42",
            },
        )
        db.add_expense(
            self.conn,
            {
                "date": "2026-05-29",
                "vendor": "Cloud Vendor",
                "paid_by": "Aditya",
                "amount": "42",
            },
        )
        db.add_payroll_entry(
            self.conn,
            {
                "month": "2026-05",
                "first_name": "Alex",
                "last_name": "Rao",
                "client": "Acme Client",
                "amount": "42",
                "gross": "100",
                "tax": "10",
                "tax_breakdown": '[{"label":"Federal Income Tax","amount":10,"total":true}]',
            },
        )
        bank_html = render_bank(self.conn)
        expenses_html = render_expenses(self.conn)
        payroll_html = render_payroll(self.conn)
        self.assertIn('name="attachment" form="bank-create-form"', bank_html)
        self.assertIn('data-extract-kind="bank"', bank_html)
        self.assertIn('data-extract-url="/bank/extract-inline"', bank_html)
        self.assertIn('id="bank-create-row"', bank_html)
        self.assertIn('data-bank-form', bank_html)
        self.assertNotIn("Smart Transaction Import", bank_html)
        self.assertNotIn("New Bank Transaction", bank_html)
        self.assertIn('name="source"', bank_html)
        self.assertIn("Filter source", bank_html)
        self.assertIn("Vamsi - $42.00", bank_html)
        self.assertIn('/bank/delete', bank_html)
        self.assertIn('data-inline-edit-toggle', bank_html)
        self.assertIn('id="bank-row-form-', bank_html)
        self.assertIn('action="/bank/update"', bank_html)
        self.assertIn('class="cell-editor"', bank_html)
        self.assertIn('inline-save-button', bank_html)
        self.assertIn('inline-save-button" type="submit"', bank_html)
        self.assertIn('data-inline-edit-cancel', bank_html)
        self.assertIn('title="Save" hidden', bank_html)
        self.assertIn('title="Cancel" hidden', bank_html)
        self.assertNotIn('/bank/edit?id=', bank_html)
        self.assertIn('href="/bank?sort=amount&direction=asc"', bank_html)
        self.assertIn('href="/bank?sort=signed&direction=asc"', bank_html)
        self.assertIn('name="attachment" form="expense-create-form"', expenses_html)
        self.assertIn('data-extract-kind="expense"', expenses_html)
        self.assertIn('data-extract-url="/expenses/extract-inline"', expenses_html)
        self.assertIn('id="expense-create-row"', expenses_html)
        self.assertNotIn("New Business Expense", expenses_html)
        self.assertIn('name="paid_by"', expenses_html)
        self.assertIn("Filter paid by", expenses_html)
        self.assertIn("Aditya - $42.00", expenses_html)
        self.assertIn('/expenses/delete', expenses_html)
        self.assertIn('id="expense-row-form-', expenses_html)
        self.assertIn('action="/expenses/update"', expenses_html)
        self.assertIn('class="cell-editor"', expenses_html)
        self.assertIn('inline-save-button', expenses_html)
        self.assertNotIn('/expenses/edit?id=', expenses_html)
        self.assertIn('href="/expenses?sort=vendor&direction=asc"', expenses_html)
        self.assertIn('href="/expenses?sort=amount&direction=asc"', expenses_html)
        self.assertIn('name="attachment" form="payroll-create-form"', payroll_html)
        self.assertIn('data-payroll-inline-file', payroll_html)
        self.assertIn('data-payroll-inline-status', payroll_html)
        self.assertIn('name="attachment_path"', payroll_html)
        self.assertNotIn("Read Paystub Files", payroll_html)
        self.assertIn("Paystub Sent", payroll_html)
        self.assertNotIn("New Payroll Entry", payroll_html)
        self.assertIn('data-inline-create-toggle', payroll_html)
        self.assertIn('id="payroll-create-row"', payroll_html)
        self.assertIn('class="inline-create-row"', payroll_html)
        self.assertIn('id="payroll-create-form"', payroll_html)
        self.assertIn('data-inline-create-cancel', payroll_html)
        self.assertIn('href="/payroll?sort=name&direction=asc"', payroll_html)
        self.assertIn('href="/payroll?sort=gross&direction=asc"', payroll_html)
        self.assertIn("Bulk Payrun Import", payroll_html)
        self.assertIn("Read payroll file(s)", payroll_html)
        self.assertIn("Vendor Pay", payroll_html)
        self.assertIn("Tax", payroll_html)
        self.assertIn("Net Pay", payroll_html)
        self.assertIn("tax-info", payroll_html)
        self.assertIn("Federal Income Tax", payroll_html)
        self.assertNotIn("Employee Pay", payroll_html)
        self.assertIn('id="payroll-row-form-', payroll_html)
        self.assertIn('action="/payroll/update"', payroll_html)
        self.assertIn('name="first_name"', payroll_html)
        self.assertIn('name="last_name"', payroll_html)
        self.assertIn('inline-save-button', payroll_html)
        self.assertIn('title="Save" hidden', payroll_html)
        self.assertIn('data-inline-edit-cancel', payroll_html)
        self.assertIn("Candidate Name", payroll_html)
        self.assertIn('name="candidate"', payroll_html)
        self.assertIn("All candidates", payroll_html)
        self.assertNotIn("All clients", payroll_html)
        self.assertIn('name="first_name"', payroll_html)
        self.assertIn('name="last_name"', payroll_html)
        self.assertIn('name="client"', payroll_html)
        self.assertIn('list="first_name-options"', payroll_html)
        self.assertIn('list="last_name-options"', payroll_html)
        self.assertIn('list="client-options"', payroll_html)
        self.assertIn("Alex", payroll_html)
        self.assertIn("Rao", payroll_html)
        self.assertIn("Acme Client", payroll_html)
        self.assertIn('/payroll/delete', payroll_html)

    def test_paystub_parser_review_and_bulk_save(self):
        text = """
        DATAFOLDIT LLC
        Name
        Ajitha Bodapothula
        Payment Date
        03 Jun 2026
        Pay Period
        01 Apr 2026 - 30 Apr 2026
        Primary Job Role Regular Pay
        $45.50 Per Hour
        176 hr
        $8,008.00
        Total Gross Pay
        Net Pay (Total Gross Pay - Total Deduction)
        $6,555.50
        """
        parsed = parse_paystub_text(text)
        self.assertEqual(parsed["month"], "2026-04")
        self.assertEqual(parsed["first_name"], "Ajitha")
        self.assertEqual(parsed["last_name"], "Bodapothula")
        self.assertEqual(parsed["credit_date"], "2026-06-03")
        self.assertAlmostEqual(parsed["vendor_pay"], 45.50, places=2)
        self.assertAlmostEqual(parsed["hours"], 176.0, places=2)
        self.assertAlmostEqual(parsed["gross"], 8008.0, places=2)
        self.assertAlmostEqual(parsed["tax"], 1452.50, places=2)
        self.assertAlmostEqual(parsed["employee_pay"], 6555.50, places=2)
        self.assertEqual(parsed["paystub_sent"], "Y")
        parsed["attachment_path"] = "/tmp/paystub.pdf"
        html = render_paystub_review(self.conn, parsed)
        self.assertIn("Review Imported Paystub", html)
        self.assertIn("Ajitha", html)
        self.assertIn("Tax", html)
        saved = add_payroll_batch(
            self.conn,
            {
                "row_count": ["1"],
                "include_0": ["on"],
                "month_0": [parsed["month"]],
                "first_name_0": [parsed["first_name"]],
                "last_name_0": [parsed["last_name"]],
                "vendor_0": [parsed["vendor"]],
                "client_0": [parsed["client"]],
                "job_start_0": [parsed["job_start"]],
                "job_end_0": [parsed["job_end"]],
                "vendor_pay_0": [str(parsed["vendor_pay"])],
                "pct_0": [str(parsed["pct"])],
                "hours_0": [str(parsed["hours"])],
                "gross_0": [str(parsed["gross"])],
                "tax_0": [str(parsed["tax"])],
                "commission_0": [str(parsed["commission"])],
                "employee_pay_0": [str(parsed["employee_pay"])],
                "credit_date_0": [parsed["credit_date"]],
                "attachment_path_0": [parsed["attachment_path"]],
                "paystub_sent_0": ["Y"],
            },
        )
        self.assertEqual(saved, 1)
        row = self.conn.execute("SELECT first_name, paystub_sent, gross, tax FROM payroll_entries WHERE first_name = 'Ajitha'").fetchone()
        self.assertEqual(row["paystub_sent"], "Y")
        self.assertAlmostEqual(row["gross"], 8008.0, places=2)
        self.assertAlmostEqual(row["tax"], 1452.50, places=2)

    def test_payrun_xls_bulk_import_rows(self):
        payrun_path = Path("/Users/vamsikrishnabhashyam/Downloads/Employee_Payrun_Summary.xls")
        if not payrun_path.exists():
            self.skipTest(f"Missing payrun workbook: {payrun_path}")
        try:
            rows = extract_paystub_rows_from_file(payrun_path)
        except RuntimeError as exc:
            self.skipTest(str(exc))
        self.assertGreaterEqual(len(rows), 2)
        rama = next(row for row in rows if row["first_name"].upper().startswith("RAMA"))
        ajitha = next(row for row in rows if row["first_name"].upper().startswith("AJITHA"))
        self.assertEqual(rama["client"], "AMEX")
        self.assertEqual(rama["vendor"], "QUANTUM CORE TECHNOLOGIES")
        self.assertAlmostEqual(rama["vendor_pay"], 38.50, places=2)
        self.assertAlmostEqual(rama["hours"], 168.0, places=2)
        self.assertAlmostEqual(rama["gross"], 6468.0, places=2)
        self.assertAlmostEqual(rama["tax"], 816.49, places=2)
        self.assertIn("Federal Income Tax", rama["tax_breakdown"])
        self.assertIn("Arizona State Tax", rama["tax_breakdown"])
        self.assertAlmostEqual(rama["employee_pay"], 5651.51, places=2)
        self.assertEqual(ajitha["client"], "HHSC")
        self.assertAlmostEqual(ajitha["vendor_pay"], 45.50, places=2)
        self.assertAlmostEqual(ajitha["hours"], 145.5, places=2)
        self.assertAlmostEqual(ajitha["gross"], 6620.25, places=2)
        self.assertAlmostEqual(ajitha["tax"], 1119.44, places=2)
        self.assertAlmostEqual(ajitha["employee_pay"], 5500.81, places=2)
        html = render_paystub_bulk_review(self.conn, rows)
        self.assertIn("Bulk Paystub Review", html)
        self.assertIn("Tax", html)
        self.assertIn("Net Pay", html)
        self.assertIn("816.49", html)

    def test_bulk_bank_review_and_save_selected_rows(self):
        html = render_transaction_bulk_review(
            self.conn,
            [
                {
                    "_ok": True,
                    "_source_name": "bank-one.csv",
                    "attachment_path": "/tmp/bank-one.csv",
                    "date": "2026-05-29",
                    "type": "Expense",
                    "category": "Software",
                    "detail": "SaaS",
                    "source": "Vamsi",
                    "amount": 120,
                    "notes": "Imported",
                },
                {
                    "_ok": False,
                    "_source_name": "bad-file.bin",
                    "attachment_path": "/tmp/bad-file.bin",
                    "error": "Could not read enough text",
                },
            ],
        )
        self.assertIn('action="/bank/create-bulk"', html)
        self.assertIn('name="row_count" value="1"', html)
        saved = add_bank_transaction_batch(
            self.conn,
            {
                "row_count": ["1"],
                "include_0": ["on"],
                "date_0": ["2026-05-29"],
                "type_0": ["Expense"],
                "category_0": ["Software"],
                "detail_0": ["SaaS"],
                "source_0": ["Vamsi"],
                "amount_0": ["120"],
                "notes_0": ["Imported"],
                "attachment_path_0": ["/tmp/bank-one.csv"],
            },
        )
        self.assertEqual(saved, 1)
        row = self.conn.execute("SELECT detail, source, attachment_path FROM bank_transactions WHERE detail = 'SaaS'").fetchone()
        self.assertEqual(row["source"], "Vamsi")
        self.assertEqual(row["attachment_path"], "/tmp/bank-one.csv")

    def test_bulk_invoice_review_and_save_selected_rows(self):
        html = render_invoice_bulk_review(
            self.conn,
            [
                {
                    "_ok": True,
                    "_source_name": "invoice-a.csv",
                    "source_pdf": "/tmp/invoice-a.csv",
                    "date": "2026-05-29",
                    "invoice_number": "INV-BULK-001",
                    "customer": "Bulk Client LLC",
                    "amount": 1250,
                    "balance_due": 1250,
                    "status": "Open",
                },
                {
                    "_ok": False,
                    "_source_name": "bad-file.bin",
                    "source_pdf": "/tmp/bad-file.bin",
                    "error": "Could not read enough text",
                },
            ],
        )
        self.assertIn('action="/invoices/create-bulk"', html)
        self.assertIn('name="row_count" value="1"', html)
        self.assertIn("Could not read this file.", html)
        saved = add_invoice_batch(
            self.conn,
            {
                "row_count": ["2"],
                "include_0": ["on"],
                "date_0": ["2026-05-29"],
                "invoice_number_0": ["INV-BULK-001"],
                "customer_0": ["Bulk Client LLC"],
                "amount_0": ["1250"],
                "commission_pct_0": ["30"],
                "commission_amount_0": ["375"],
                "due_date_0": ["2026-06-29"],
                "status_0": ["Not Received"],
                "balance_due_0": ["1250"],
                "source_pdf_0": ["/tmp/invoice-a.csv"],
                "date_1": ["2026-05-29"],
                "invoice_number_1": ["INV-BULK-002"],
                "customer_1": ["Skipped LLC"],
                "amount_1": ["500"],
            },
        )
        self.assertEqual(saved, 1)
        rows = self.conn.execute("SELECT invoice_number, source_pdf FROM invoices WHERE invoice_number LIKE 'INV-BULK-%'").fetchall()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["invoice_number"], "INV-BULK-001")
        self.assertEqual(rows[0]["source_pdf"], "/tmp/invoice-a.csv")


if __name__ == "__main__":
    unittest.main()
