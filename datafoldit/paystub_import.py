from __future__ import annotations

import json
import re
import shutil
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from . import db
from .transaction_import import extract_text as extract_attachment_text


MONEY_RE = re.compile(r"\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{2})|[0-9]+(?:\.[0-9]{2}))")
DATE_RE = re.compile(
    r"\b\d{1,2}\s+(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\s+\d{4}\b"
    r"|\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\s+\d{1,2},?\s+\d{4}\b"
    r"|\b\d{4}-\d{2}-\d{2}\b|\b\d{1,2}/\d{1,2}/\d{2,4}\b",
    re.I,
)
PAY_PERIOD_RE = re.compile(f"({DATE_RE.pattern})\\s*-\\s*({DATE_RE.pattern})", re.I)
PAYRUN_TAX_COLUMNS = [
    ("Federal Income Tax", "federal income tax"),
    ("Federal Unemployment Tax", "federal unemployment tax"),
    ("Texas State Unemployment Tax", "texas state unemployment tax"),
    ("Arizona State Tax", "arizona state tax"),
    ("Arizona State Unemployment Tax", "arizona state unemployment tax"),
]


def extract_paystub_from_file(path: str | Path) -> dict[str, Any]:
    source = Path(path).expanduser()
    if not source.exists():
        raise FileNotFoundError(source)
    if source.suffix.lower() in {".xls", ".xlsx", ".xlsm"}:
        rows = extract_payrun_rows(source)
        if not rows:
            raise RuntimeError("No payroll rows found in the payrun workbook.")
        return rows[0]
    text = extract_paystub_text(source)
    fields = parse_paystub_text(text)
    fields["attachment_path"] = str(source)
    fields["raw_text_excerpt"] = text[:3000]
    return fields


def extract_paystub_rows_from_file(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path).expanduser()
    if not source.exists():
        raise FileNotFoundError(source)
    if source.suffix.lower() in {".xls", ".xlsx", ".xlsm"}:
        rows = extract_payrun_rows(source)
        if not rows:
            raise RuntimeError("No payroll rows found in the payrun workbook.")
        return rows
    return [extract_paystub_from_file(source)]


def extract_paystub_text(path: Path) -> str:
    text = pdftotext(path) if path.suffix.lower() == ".pdf" else ""
    if not useful_text(text):
        text = extract_attachment_text(path)
    if not useful_text(text):
        raise RuntimeError("Could not read enough paystub text from the uploaded file.")
    return text


def extract_payrun_rows(path: Path) -> list[dict[str, Any]]:
    workbook_path = path
    temp_dir: tempfile.TemporaryDirectory[str] | None = None
    if path.suffix.lower() == ".xls":
        temp_dir = tempfile.TemporaryDirectory(prefix="datafoldit-payrun-")
        try:
            workbook_path = convert_xls_to_xlsx(path, Path(temp_dir.name))
        except Exception:
            if temp_dir:
                temp_dir.cleanup()
            raise
    try:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            sheet = workbook.worksheets[0]
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
    finally:
        if temp_dir:
            temp_dir.cleanup()
    if not rows:
        return []
    headers = [normalize_header(value) for value in rows[0]]
    extracted: list[dict[str, Any]] = []
    for values in rows[1:]:
        item = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
        if not item.get("employee name") or not item.get("payment date"):
            continue
        parsed = parse_payrun_row(item, path)
        if parsed:
            extracted.append(parsed)
    return extracted


def convert_xls_to_xlsx(path: Path, outdir: Path) -> Path:
    soffice = find_tool("soffice")
    if not soffice:
        raise RuntimeError("Could not read old .xls payrun files because LibreOffice/soffice is not installed.")
    profile_dir = outdir / "lo-profile"
    result = subprocess.run(
        [
            soffice,
            "--headless",
            f"-env:UserInstallation=file://{profile_dir}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(outdir),
            str(path),
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=30,
    )
    if result.returncode != 0:
        raise RuntimeError("Could not convert .xls payrun file to .xlsx for import.")
    converted = outdir / f"{path.stem}.xlsx"
    if not converted.exists():
        matches = sorted(outdir.glob("*.xlsx"))
        if not matches:
            raise RuntimeError("Converted payrun workbook was not found.")
        converted = matches[0]
    return converted


def parse_payrun_row(item: dict[str, Any], source: Path) -> dict[str, Any]:
    full_name = str(item.get("employee name") or "").strip()
    first_name, last_name = split_name(full_name)
    payment_date = db.normalize_date(item.get("payment date"))
    period_start = db.normalize_date(item.get("payperiod start"))
    period_end = db.normalize_date(item.get("payperiod end"))
    hours = parse_payrun_hours(item.get("regular pay - primary job role hours"))
    gross = db.amount_value(item.get("total earnings") or item.get("regular pay - primary job role"))
    vendor_pay = round(gross / hours, 2) if gross and hours else 0.0
    tax = db.amount_value(item.get("total deductions"))
    tax_breakdown = payrun_tax_breakdown(item, tax)
    client, vendor = infer_client_vendor(item.get("work location"))
    return {
        "month": (period_end or period_start or payment_date or "")[:7],
        "first_name": first_name,
        "last_name": last_name,
        "vendor": vendor,
        "client": client,
        "job_start": period_start,
        "job_end": period_end,
        "vendor_pay": vendor_pay,
        "pct": 0,
        "hours": hours,
        "gross": gross,
        "tax": tax,
        "tax_breakdown": tax_breakdown,
        "commission": 0,
        "employee_pay": db.amount_value(item.get("net pay")),
        "credit_date": payment_date,
        "attachment_path": str(source),
        "paystub_sent": "Y",
    }


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def payrun_tax_breakdown(item: dict[str, Any], total_tax: float) -> str:
    lines: list[dict[str, float | str]] = []
    for label, key in PAYRUN_TAX_COLUMNS:
        amount = db.amount_value(item.get(key))
        if amount:
            lines.append({"label": label, "amount": round(amount, 2)})
    if total_tax:
        lines.append({"label": "Total Deductions", "amount": round(total_tax, 2), "total": True})
    return json.dumps(lines, separators=(",", ":")) if lines else ""


def parse_payrun_hours(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    match = re.fullmatch(r"(\d+):(\d{1,2})", text)
    if match:
        return round(float(match.group(1)) + float(match.group(2)) / 60, 4)
    return db.amount_value(text)


def infer_client_vendor(value: Any) -> tuple[str, str]:
    text = str(value or "").strip()
    lower = text.lower()
    if "american express" in lower or "amex" in lower:
        return "AMEX", "QUANTUM CORE TECHNOLOGIES"
    if "health and human services" in lower or "hhsc" in lower:
        return "HHSC", "SRB SYSTEMS"
    return "", ""


def pdftotext(path: Path) -> str:
    tool = find_tool("pdftotext")
    if not tool:
        return ""
    try:
        result = subprocess.run(
            [tool, str(path), "-"],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            text=True,
            timeout=20,
        )
        return result.stdout
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired):
        return ""


def find_tool(name: str) -> str | None:
    found = shutil.which(name)
    if found:
        return found
    bundled_bin = Path.home() / ".cache/codex-runtimes/codex-primary-runtime/dependencies/bin"
    for directory in (
        "/opt/homebrew/bin",
        "/usr/local/bin",
        "/usr/bin",
        str(bundled_bin),
        "/Applications/LibreOffice.app/Contents/MacOS",
    ):
        candidate = Path(directory) / name
        if candidate.exists():
            return str(candidate)
    return None


def useful_text(text: str) -> bool:
    return len(re.findall(r"[A-Za-z0-9]{2,}", text or "")) >= 8


def parse_paystub_text(text: str) -> dict[str, Any]:
    normalized = normalize_text(text)
    lines = normalized.splitlines()
    full_name = value_after_label(lines, "Name") or ""
    first_name, last_name = split_name(full_name)
    payment_date = parse_date(value_after_label(lines, "Payment Date") or "")
    period_start, period_end = extract_pay_period(normalized)
    rate = extract_rate(normalized)
    hours = extract_hours(normalized)
    gross = extract_regular_pay_amount(lines) or money_after_label(lines, "Total Gross Pay") or 0.0
    net_pay = money_after_label(lines, "Net Pay") or money_after_label(lines, "YOUR NET PAY IS") or 0.0
    deductions = round(gross - net_pay, 2) if gross and net_pay and gross >= net_pay else 0.0
    tax_breakdown = (
        json.dumps([{"label": "Total Deductions", "amount": deductions, "total": True}], separators=(",", ":"))
        if deductions
        else ""
    )
    return {
        "month": (period_end or period_start or payment_date or "")[:7],
        "first_name": first_name,
        "last_name": last_name,
        "vendor": extract_employer(lines),
        "client": "",
        "job_start": period_start,
        "job_end": period_end,
        "vendor_pay": rate,
        "pct": 0,
        "hours": hours,
        "gross": gross,
        "tax": deductions,
        "tax_breakdown": tax_breakdown,
        "commission": 0,
        "employee_pay": net_pay,
        "credit_date": payment_date,
        "paystub_sent": "Y",
    }


def normalize_text(text: str) -> str:
    lines = [re.sub(r"\s+", " ", line).strip() for line in (text or "").splitlines()]
    return "\n".join(line for line in lines if line)


def value_after_label(lines: list[str], label: str) -> str | None:
    label_lower = label.lower()
    for index, line in enumerate(lines):
        lower = line.lower()
        if label_lower not in lower:
            continue
        remainder = re.sub(re.escape(label), "", line, flags=re.I).strip(" :")
        if remainder:
            return remainder
        for following in lines[index + 1 : index + 5]:
            if following:
                return following
    return None


def money_after_label(lines: list[str], label: str) -> float | None:
    label_lower = label.lower()
    for index, line in enumerate(lines):
        if label_lower not in line.lower():
            continue
        for candidate in [line, *lines[index + 1 : index + 8]]:
            value = first_money(candidate)
            if value is not None:
                return value
    return None


def extract_regular_pay_amount(lines: list[str]) -> float | None:
    for index, line in enumerate(lines):
        if "regular pay" not in line.lower():
            continue
        amounts = [first_money(candidate) for candidate in lines[index : index + 6]]
        real_amounts = [value for value in amounts if value and value > 100]
        if real_amounts:
            return real_amounts[0]
    return None


def first_money(text: str) -> float | None:
    match = MONEY_RE.search(text or "")
    return db.amount_value(match.group(1)) if match else None


def extract_rate(text: str) -> float:
    match = re.search(r"\$?\s*([0-9,]+(?:\.[0-9]{2})?)\s*Per\s+Hour", text, re.I)
    return db.amount_value(match.group(1)) if match else 0.0


def extract_hours(text: str) -> float:
    match = re.search(r"\b([0-9]+(?:\.[0-9]+)?)\s*hr\b", text, re.I)
    return db.amount_value(match.group(1)) if match else 0.0


def extract_pay_period(text: str) -> tuple[str | None, str | None]:
    match = PAY_PERIOD_RE.search(text)
    if not match:
        return None, None
    return parse_date(match.group(1)), parse_date(match.group(2))


def parse_date(value: str) -> str | None:
    value = value.strip()
    for fmt in ("%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y", "%b %d %Y", "%B %d %Y"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return db.normalize_date(value)


def split_name(name: str) -> tuple[str, str]:
    parts = [part for part in re.split(r"\s+", name.strip()) if part]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def extract_employer(lines: list[str]) -> str:
    for line in lines[:5]:
        if "datafoldit" in line.lower():
            return line
    return ""
