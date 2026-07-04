from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from . import db


DEFAULT_SOURCE_XLSX = Path("/Users/vamsikrishnabhashyam/Downloads/company_expenses.xlsx")


def import_company_workbook(
    conn,
    workbook_path: str | Path = DEFAULT_SOURCE_XLSX,
    replace: bool = False,
) -> dict[str, int]:
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, data_only=True)
    db.init_db(conn)
    if replace:
        db.clear_operational_data(conn)

    counts = {
        "expenses": import_expenses(conn, workbook),
        "bank_transactions": import_bank_transactions(conn, workbook),
        "payroll_entries": import_payroll(conn, workbook),
        "invoices": import_invoices(conn, workbook),
    }
    db.set_setting(conn, "last_import_path", str(path))
    db.set_setting(conn, "last_import_at", datetime.now(UTC).isoformat(timespec="seconds"))
    conn.commit()
    return counts


def rows_from_table(workbook, sheet_name: str, table_name: str) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [
        clean_header(ws.cell(row=min_row, column=col).value)
        for col in range(min_col, max_col + 1)
    ]
    rows: list[dict[str, Any]] = []
    for row_number in range(min_row + 1, max_row + 1):
        values = [ws.cell(row=row_number, column=col).value for col in range(min_col, max_col + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({headers[index]: values[index] for index in range(len(headers))})
    return rows


def clean_header(value: Any) -> str:
    return str(value or "").strip()


def value_from(item: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in item and item.get(key) not in (None, ""):
            return item.get(key)
    return None


def import_expenses(conn, workbook) -> int:
    rows = []
    for item in rows_from_table(workbook, "Expenses", "Log"):
        if item.get("Date") is None or item.get("Amount") in (None, ""):
            continue
        rows.append(
            {
                "date": item.get("Date"),
                "category": item.get("Category"),
                "vendor": item.get("Vendor"),
                "description": item.get("Description"),
                "amount": item.get("Amount"),
                "paid_by": item.get("Paid By"),
                "frequency": item.get("Frequency"),
                "notes": item.get("Notes"),
            }
        )
    return db.bulk_insert(conn, "expenses", rows)


def import_bank_transactions(conn, workbook) -> int:
    rows = []
    for item in rows_from_table(workbook, "Account Summary", "Txns"):
        if item.get("Date") is None or item.get("Amount") in (None, ""):
            continue
        rows.append(
            {
                "date": item.get("Date"),
                "month": month_to_key(item.get("Date"), item.get("Month")),
                "type": normalize_bank_type(item.get("Type")),
                "category": item.get("Category"),
                "detail": item.get("Vendor / Detail"),
                "source": item.get("Paid By / Source"),
                "amount": item.get("Amount"),
            }
        )
    return db.bulk_insert(conn, "bank_transactions", rows)


def import_payroll(conn, workbook) -> int:
    rows = []
    for item in rows_from_table(workbook, "Payroll", "Payroll"):
        if item.get("Month") is None or not any(item.get(key) for key in ("Gross", "Hours", "First Name")):
            continue
        rows.append(
            {
                "month": item.get("Month"),
                "first_name": item.get("First Name"),
                "last_name": item.get("Last Name"),
                "vendor": item.get("Vendor"),
                "client": item.get("Client"),
                "job_start": item.get("Job Start"),
                "job_end": item.get("Job End"),
                "vendor_pay": value_from(item, "Vendor Pay", "Pay Rate / Hour"),
                "pct": value_from(item, "Pct", "Commission %"),
                "hours": item.get("Hours"),
                "gross": value_from(item, "Gross", "Total Earnings"),
                "tax": value_from(item, "Tax", "Total Deductions"),
                "commission": value_from(item, "Commission", "Commission Amount"),
                "employee_pay": value_from(item, "Employee Pay", "Net Pay"),
                "credit_date": value_from(item, "Credit Date", "Payment Date"),
            }
        )
    return db.bulk_insert(conn, "payroll_entries", rows)


def import_invoices(conn, workbook) -> int:
    rows = []
    for item in rows_from_table(workbook, "Invoices", "Inv"):
        if item.get("Date") is None or not item.get("Invoice #"):
            continue
        rows.append(
            {
                "date": item.get("Date"),
                "invoice_number": item.get("Invoice #"),
                "customer": item.get("Customer"),
                "is_void": item.get("Void?"),
                "received": item.get("Received"),
                "due_date": item.get("Due Date"),
                "amount": item.get("Amount"),
                "commission_pct": value_from(item, "Commission %", "Commission Percent", "Pct"),
                "commission_amount": value_from(item, "Commission Amount"),
                "status": item.get("Status"),
                "balance_due": item.get("Balance Due"),
            }
        )
    return db.bulk_insert(conn, "invoices", rows)


def month_to_key(date_value: Any, month_value: Any) -> str | None:
    normalized = db.normalize_month(date_value)
    if normalized:
        return normalized
    normalized = db.normalize_month(month_value)
    if normalized:
        return normalized
    return db.clean_text(month_value)


def normalize_bank_type(value: Any) -> str | None:
    text = db.clean_text(value)
    if text is None:
        return None
    canonical = {
        "opening": "Opening",
        "deposit": "Deposit",
        "transfer in": "Transfer In",
        "adjustment in": "Adjustment In",
        "expense": "Expense",
        "withdrawal": "Withdrawal",
        "transfer out": "Transfer Out",
        "adjustment out": "Adjustment Out",
    }
    return canonical.get(text.lower(), text)


def export_report_workbook(
    conn,
    output_path: str | Path,
    period: str = "all",
    month: str | None = None,
    day: str | None = None,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    report_data = db.fetch_report_data(conn, period=period, month=month, day=day)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    build_summary_sheet(summary, report_data, period, month, day)
    build_bank_sheet(workbook.create_sheet("Bank Transactions"), report_data["bank"])
    build_expenses_sheet(workbook.create_sheet("Expenses"), report_data["expenses"])
    build_payroll_sheet(workbook.create_sheet("Payroll"), report_data["payroll"])
    build_invoices_sheet(workbook.create_sheet("Invoices"), report_data["invoices"])
    for ws in workbook.worksheets:
        style_sheet(ws)
    workbook.save(output)
    return output


def build_summary_sheet(ws, report_data, period: str, month: str | None, day: str | None) -> None:
    metrics = report_data["metrics"]
    label = "All Time"
    if period == "monthly" and month:
        label = month
    elif period == "daily" and day:
        label = day
    rows = [
        ["DataFold IT Operations Report", "", "", ""],
        ["Report Period", label, "Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        [],
        ["Metric", "Amount", "Metric", "Amount"],
        ["Current Bank Balance", metrics["current_balance"], "Business Expenses", metrics["total_expenses"]],
        ["Invoice Total", metrics["invoice_total"], "Invoice Outstanding", metrics["invoice_outstanding"]],
        ["Invoice Paid", metrics["invoice_paid"], "Commission Received", metrics["invoice_commission_received"]],
        ["Payroll Gross", metrics["payroll_gross"], "Payroll Tax", metrics["payroll_tax"]],
        [],
        ["Record Counts", "", "", ""],
        ["Bank Transactions", len(report_data["bank"]), "Expenses", len(report_data["expenses"])],
        ["Payroll Entries", len(report_data["payroll"]), "Invoices", len(report_data["invoices"])],
    ]
    for row in rows:
        ws.append(row)
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E5F")
    ws["A4"].font = ws["C4"].font = Font(bold=True)
    ws["A10"].font = Font(bold=True)


def build_bank_sheet(ws, rows) -> None:
    ws.append(["Date", "Month", "Type", "Category", "Detail", "Source", "Amount", "Signed Amount", "Attachment"])
    for row in rows:
        ws.append(
            [
                row["date"],
                row["month"],
                row["type"],
                row["category"],
                row["detail"],
                row["source"],
                row["amount"],
                db.bank_signed_amount(row),
                row["attachment_path"] if "attachment_path" in row.keys() else None,
            ]
        )


def build_expenses_sheet(ws, rows) -> None:
    ws.append(["Date", "Category", "Vendor", "Description", "Amount", "Paid By", "Frequency", "Notes", "Attachment"])
    for row in rows:
        ws.append(
            [
                row["date"],
                row["category"],
                row["vendor"],
                row["description"],
                row["amount"],
                row["paid_by"],
                row["frequency"],
                row["notes"],
                row["attachment_path"] if "attachment_path" in row.keys() else None,
            ]
        )


def build_payroll_sheet(ws, rows) -> None:
    ws.append(
        [
            "Month",
            "First Name",
            "Last Name",
            "Vendor",
            "Client",
            "Job Start",
            "Job End",
            "Vendor Pay",
            "Hours",
            "Gross",
            "Tax",
            "Credit Date",
            "Paystub Sent",
            "Attachment",
        ]
    )
    for row in rows:
        ws.append(
            [
                row["month"],
                row["first_name"],
                row["last_name"],
                row["vendor"],
                row["client"],
                row["job_start"],
                row["job_end"],
                row["vendor_pay"],
                row["hours"],
                row["gross"],
                row["tax"] if "tax" in row.keys() else 0,
                row["credit_date"],
                row["paystub_sent"] if "paystub_sent" in row.keys() else None,
                row["attachment_path"] if "attachment_path" in row.keys() else None,
            ]
        )


def build_invoices_sheet(ws, rows) -> None:
    ws.append(
        [
            "Date",
            "Invoice #",
            "Customer",
            "Void?",
            "Received",
            "Due Date",
            "Amount",
            "Commission %",
            "Commission Amount",
            "Status",
            "Balance Due",
            "Source PDF",
        ]
    )
    for row in rows:
        ws.append(
            [
                row["date"],
                row["invoice_number"],
                row["customer"],
                "Y" if row["is_void"] else "",
                row["received"],
                row["due_date"],
                row["amount"],
                db.commission_fraction(row["commission_pct"] if "commission_pct" in row.keys() else 30),
                row["commission_amount"] if "commission_amount" in row.keys() else 0,
                row["status"],
                row["balance_due"],
                row["source_pdf"] if "source_pdf" in row.keys() else None,
            ]
        )


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="234E52")
    header_font = Font(bold=True, color="FFFFFF")
    currency_columns = {
        "Amount",
        "Signed Amount",
        "Vendor Pay",
        "Gross",
        "Tax",
        "Commission",
        "Commission Amount",
        "Employee Pay",
        "Balance Due",
    }
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    if ws.max_row >= 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    headers = [cell.value for cell in ws[1]]
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        max_len = max(
            len(str(ws.cell(row=row, column=index).value or ""))
            for row in range(1, min(ws.max_row, 200) + 1)
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 34)
        if header in currency_columns:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=index).number_format = '$#,##0.00'
        if header in {"Pct", "Commission %"}:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=index).number_format = "0.0%"
