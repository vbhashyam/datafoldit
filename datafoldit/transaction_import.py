from __future__ import annotations

import re
import shutil
import subprocess
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook

from . import db


DATE_RE = re.compile(
    r"\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|"
    r"Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\s+\d{1,2},?\s+\d{4}\b"
    r"|\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b"
    r"|\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b",
    re.I,
)
MONEY_RE = re.compile(
    r"(?<![\w/])[-(]?\s*\$?\s*(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d{2})?\)?(?![\w/])"
)

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff", ".bmp", ".heic"}
TEXT_EXTS = {".txt", ".csv", ".tsv"}
BALANCE_WORDS = ("balance", "available", "remaining", "ending")
AMOUNT_WORDS = (
    "amount",
    "total",
    "payment",
    "paid",
    "charge",
    "debit",
    "credit",
    "deposit",
    "withdrawal",
    "sale",
    "purchase",
)


def extract_transaction_from_file(path: str | Path) -> dict[str, Any]:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(source)
    text = extract_text(source)
    fields = parse_transaction_text(text, source.name)
    fields["attachment_path"] = str(source)
    fields["raw_text_excerpt"] = text[:3000]
    fields["source_filename"] = source.name
    return fields


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        text = extract_xlsx_text(path)
    elif suffix == ".docx":
        text = extract_docx_text(path)
    elif suffix in TEXT_EXTS:
        text = path.read_text(errors="replace")
    elif suffix in IMAGE_EXTS:
        text = ocr_image(path)
    else:
        text = ""

    if not useful_text(text):
        text = document_ocr(path)
    if not useful_text(text):
        raise RuntimeError("Could not read enough text from the attachment. Try a clearer screenshot or PDF.")
    return text


def extract_xlsx_text(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    try:
        for sheet in workbook.worksheets[:3]:
            lines.append(sheet.title)
            for row in sheet.iter_rows(max_row=120, max_col=24, values_only=True):
                values = [format_cell(value) for value in row if value not in (None, "")]
                if values:
                    lines.append(" | ".join(values))
                if len(lines) >= 250:
                    break
    finally:
        workbook.close()
    return "\n".join(lines)


def extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    words = [node.text for node in root.iter(f"{namespace}t") if node.text]
    return "\n".join(words)


def ocr_image(path: Path) -> str:
    tesseract = shutil.which("tesseract")
    if not tesseract:
        return ""
    try:
        result = subprocess.run(
            [tesseract, str(path), "stdout"],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            text=True,
        )
        return result.stdout
    except subprocess.CalledProcessError:
        return ""


def document_ocr(path: Path) -> str:
    if path.suffix.lower() == ".pdf":
        text = pdftoppm_ocr(path)
        if useful_text(text):
            return text
    return quicklook_ocr(path)


def pdftoppm_ocr(path: Path) -> str:
    pdftoppm = shutil.which("pdftoppm")
    tesseract = shutil.which("tesseract")
    if not pdftoppm or not tesseract:
        return ""
    with tempfile.TemporaryDirectory(prefix="datafoldit-attachment-ocr-") as tmp:
        image_prefix = Path(tmp) / "page"
        try:
            subprocess.run(
                [pdftoppm, "-png", "-singlefile", "-r", "220", str(path), str(image_prefix)],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            result = subprocess.run(
                [tesseract, str(image_prefix.with_suffix(".png")), "stdout"],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                text=True,
            )
            return result.stdout
        except subprocess.CalledProcessError:
            return ""


def quicklook_ocr(path: Path) -> str:
    qlmanage = shutil.which("qlmanage")
    tesseract = shutil.which("tesseract")
    if not qlmanage or not tesseract:
        return ""
    with tempfile.TemporaryDirectory(prefix="datafoldit-attachment-ocr-") as tmp:
        tmp_path = Path(tmp)
        try:
            subprocess.run(
                [qlmanage, "-t", "-s", "1800", "-o", str(tmp_path), str(path)],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
        except subprocess.CalledProcessError:
            return ""
        images = sorted(tmp_path.glob("*.png"))
        if not images:
            return ""
        try:
            result = subprocess.run(
                [tesseract, str(images[0]), "stdout"],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                text=True,
            )
            return result.stdout
        except subprocess.CalledProcessError:
            return ""


def useful_text(text: str) -> bool:
    words = re.findall(r"[A-Za-z0-9]{2,}", text or "")
    return len(words) >= 4


def parse_transaction_text(text: str, source_name: str | None = None) -> dict[str, Any]:
    normalized = normalize_text(text)
    amount, signed_amount = extract_amount(normalized)
    tx_type = infer_transaction_type(normalized, signed_amount)
    tx_date = extract_labeled_date(normalized) or extract_first_date(normalized)
    detail = extract_detail(normalized, source_name)
    category = infer_category(normalized, detail, tx_type)
    source = infer_source(normalized)
    confidence = confidence_score(tx_date, amount, detail, category, tx_type)
    notes = f"Smart imported from {source_name or 'attachment'}. Confidence {confidence}%."
    return {
        "date": tx_date,
        "type": tx_type,
        "category": category,
        "detail": detail,
        "source": source,
        "amount": abs(amount),
        "notes": notes,
        "confidence": confidence,
    }


def normalize_text(text: str) -> str:
    lines = [re.sub(r"\s+", " ", line).strip() for line in (text or "").splitlines()]
    return "\n".join(line for line in lines if line)


def extract_labeled_date(text: str) -> str | None:
    labels = ("transaction date", "payment date", "paid on", "posted", "date")
    for line in text.splitlines():
        lower = line.lower()
        if any(label in lower for label in labels):
            match = DATE_RE.search(line)
            if match:
                return parse_date(match.group(0))
    return None


def extract_first_date(text: str) -> str | None:
    match = DATE_RE.search(text)
    return parse_date(match.group(0)) if match else None


def parse_date(value: str) -> str | None:
    value = value.strip().replace("/", "-")
    for fmt in ("%b %d, %Y", "%B %d, %Y", "%b %d %Y", "%B %d %Y", "%m-%d-%Y", "%m-%d-%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return db.normalize_date(value)


def extract_amount(text: str) -> tuple[float, float]:
    candidates: list[tuple[int, int, float]] = []
    all_amounts: list[float] = []
    for index, line in enumerate(text.splitlines()):
        lower = line.lower()
        if any(word in lower for word in BALANCE_WORDS):
            continue
        for raw, amount in money_values(line):
            if amount == 0:
                continue
            score = 1
            if "$" in raw or "." in raw:
                score += 2
            if any(word in lower for word in AMOUNT_WORDS):
                score += 4
            if "total" in lower or "amount" in lower:
                score += 2
            candidates.append((score, index, amount))
            all_amounts.append(amount)
    if candidates:
        candidates.sort(key=lambda item: (item[0], item[1], abs(item[2])))
        amount = candidates[-1][2]
        return abs(amount), amount
    all_seen = [amount for _, amount in money_values(text)]
    if not all_seen:
        return 0.0, 0.0
    amount = max(all_seen, key=lambda value: abs(value))
    return abs(amount), amount


def money_values(text: str) -> list[tuple[str, float]]:
    values = []
    for match in MONEY_RE.finditer(text):
        raw = match.group(0).strip()
        amount = db.amount_value(raw)
        if not amount:
            continue
        plain_digits = "$" not in raw and "." not in raw and "," not in raw
        if plain_digits and 1900 <= abs(amount) <= 2100:
            continue
        values.append((raw, amount))
    return values


def infer_transaction_type(text: str, signed_amount: float) -> str:
    lower = text.lower()
    deposit_phrases = (
        "deposit",
        "ach credit",
        "payment received",
        "received from",
        "incoming",
        "direct deposit",
        "credited to",
        "paid to you",
    )
    expense_phrases = (
        "receipt",
        "purchase",
        "debit",
        "withdrawal",
        "charged",
        "payment to",
        "paid to",
        "sent to",
        "card ending",
        "visa",
        "mastercard",
        "amex",
    )
    deposit_score = sum(1 for phrase in deposit_phrases if phrase in lower)
    expense_score = sum(1 for phrase in expense_phrases if phrase in lower)
    if signed_amount < 0:
        expense_score += 2
    if deposit_score > expense_score:
        return "Deposit"
    return "Expense"


def extract_detail(text: str, source_name: str | None = None) -> str | None:
    label_re = re.compile(r"\b(?:merchant|vendor|payee|paid to|payment to|to|from|sender|description)\b\s*[:\-]?\s*(.+)", re.I)
    for line in text.splitlines():
        match = label_re.search(line)
        if match:
            candidate = clean_detail(match.group(1))
            if candidate:
                return candidate
    for line in text.splitlines()[:14]:
        candidate = clean_detail(line)
        if candidate:
            return candidate
    return Path(source_name).stem if source_name else None


def clean_detail(line: str) -> str | None:
    text = line.strip(" :-|")
    if not text or len(text) < 3 or len(text) > 90:
        return None
    lower = text.lower()
    if DATE_RE.search(text) or money_values(text):
        return None
    blocked = ("transaction", "receipt", "invoice", "date", "amount", "total", "balance", "account", "card")
    if any(word in lower for word in blocked):
        return None
    if not re.search(r"[A-Za-z]{3,}", text):
        return None
    return text


def infer_category(text: str, detail: str | None, tx_type: str) -> str:
    if tx_type == "Deposit":
        return "Client Payment"
    lower = f"{text}\n{detail or ''}".lower()
    categories = [
        ("Software", ("zoom", "microsoft", "google", "github", "aws", "openai", "slack", "adobe", "software", "subscription", "cloud")),
        ("Bank Fees", ("bank fee", "service fee", "wire fee", "monthly fee")),
        ("Payroll", ("payroll", "salary", "commission")),
        ("Travel", ("hotel", "airlines", "uber", "lyft", "travel", "flight")),
        ("Meals", ("restaurant", "cafe", "coffee", "meal", "food")),
        ("Internet", ("internet", "comcast", "verizon", "at&t", "phone")),
        ("Professional Services", ("legal", "accounting", "consulting", "services")),
        ("Taxes", ("tax", "irs", "state of")),
    ]
    for category, keywords in categories:
        if any(keyword in lower for keyword in keywords):
            return category
    return "Business Expense"


def infer_source(text: str) -> str | None:
    lower = text.lower()
    if "zelle" in lower:
        return "Zelle"
    if "paypal" in lower:
        return "PayPal"
    if "ach" in lower:
        return "ACH"
    if any(card in lower for card in ("visa", "mastercard", "amex", "card ending", "credit card", "debit card")):
        return "Card"
    if "bank" in lower or "checking" in lower:
        return "Bank"
    return None


def confidence_score(date_value: str | None, amount: float, detail: str | None, category: str | None, tx_type: str | None) -> int:
    score = 0
    if date_value:
        score += 30
    if amount:
        score += 35
    if detail:
        score += 20
    if category:
        score += 10
    if tx_type:
        score += 5
    return min(score, 95)


def format_cell(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()
